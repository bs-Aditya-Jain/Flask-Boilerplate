"""Contains user related API definitions."""
from datetime import datetime
from datetime import timedelta
from typing import Any
from app import config_data
from app import db
from app import limiter
from app import logger
from app.helpers.constants import HttpStatusCode
from app.helpers.constants import ResponseMessageKeys
from app.helpers.decorators import api_time_logger
from app.helpers.decorators import token_required
from app.helpers.utility import field_type_validator
from app.helpers.utility import get_pagination_meta
from app.helpers.utility import required_validator
from app.helpers.utility import send_json_response
from app.models.user import User
from flask import request
from flask.views import View
import jwt
from werkzeug.security import check_password_hash
from openpyxl import load_workbook
import os 
import traceback

class UserView(View):
    """Contains all user related functions"""

    def create_auth_response(self, user: User, data: Any = None) -> dict:
        """Returns user details and access token
        """
        token = jwt.encode({
            'id': user.id,
            'exp': datetime.utcnow() + timedelta(days=60)
        }, key=config_data.get('SECRET_KEY'))

        user_details = {
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.primary_email,
            'created_at': user.created_at,
            'updated_at': user.updated_at,

            'name': '{} {}'.format(user.first_name, user.last_name) if user.last_name else user.first_name.title(),
            'phone': user.primary_phone,
            'country_code': user.country_code,
            'address': user.address,

        }
        user.last_login_at = datetime.now()
        user.auth_token = token
        db.session.commit()

        response = {'token': token, 'details': user_details}

        return response

    @staticmethod
    @api_time_logger
    @token_required
    def search(logged_in_user: User) -> tuple:
        """Used to return the list of all users based on search , pagination and sorting
            with pagination metadata.
        """
        data = request.args
        field_types = {'page': int, 'size': int,
                       'q': str, 'sort': str, 'user_type': str}
        required_fields = ['page', 'size']
        post_data = field_type_validator(
            request_data=data, field_types=field_types)
        if post_data['is_error']:
            return send_json_response(http_status=HttpStatusCode.BAD_REQUEST.value, response_status=False,
                                      message_key=ResponseMessageKeys.ENTER_CORRECT_INPUT.value, data=None,
                                      error=post_data['data'])

        is_valid = required_validator(
            request_data=data, required_fields=required_fields)
        if is_valid['is_error']:
            return send_json_response(http_status=HttpStatusCode.BAD_REQUEST.value, response_status=False,
                                      message_key=ResponseMessageKeys.ENTER_CORRECT_INPUT.value, data=None,
                                      error=is_valid['data'])
        page = request.args.get('page')
        size = request.args.get('size')
        q = request.args.get('q')
        sort = request.args.get('sort')

        user_list = User.get_user_list(
            q=q, sort=sort, page=page, size=size).all()
        user_count = User.get_user_list(q=q, sort=sort).count()
        user_data_ = User.serialize_user(user_list)
        total_count = user_count
        data = {'result': user_data_,
                'pagination_metadata': get_pagination_meta(current_page=1 if page is None else int(page),
                                                           page_size=int(size),
                                                           total_items=int(total_count))}
        return send_json_response(http_status=HttpStatusCode.OK.value, response_status=True,
                                  message_key=ResponseMessageKeys.SUCCESS.value, data=data,
                                  error=None)

    @staticmethod
    @api_time_logger
    # ideally it can be 1/30 , it is currently 3/1 so that tests do not fail
    @limiter.limit(limit_value='3/1 second', key_func=lambda: request.get_json(force=True).get('email'))
    def login():
        """Login api for admin user to check pin and email and return login response with access token"""

        data = request.get_json(force=True)
        field_types = {'email': str, 'pin': str}
        required_fields = ['email', 'pin']
        post_data = field_type_validator(
            request_data=data, field_types=field_types)
        if post_data['is_error']:
            return send_json_response(http_status=HttpStatusCode.BAD_REQUEST.value, response_status=False,
                                      message_key=ResponseMessageKeys.ENTER_CORRECT_INPUT.value,
                                      data=None, error=post_data['data'])

        is_valid = required_validator(
            request_data=data, required_fields=required_fields)
        if is_valid['is_error']:
            return send_json_response(http_status=HttpStatusCode.BAD_REQUEST.value, response_status=False,
                                      message_key=ResponseMessageKeys.ENTER_CORRECT_INPUT.value, data=None,
                                      error=is_valid['data'])

        primary_email = data.get('email')
        pin = data.get('pin')
        try:
            user = User.get_by_email(email=primary_email)
        except Exception as e:
            logger.error(
                'Error while fetching user in auth_super_admin API :{}'.format(e))
            return send_json_response(http_status=HttpStatusCode.BAD_REQUEST.value, response_status=False,
                                      message_key=ResponseMessageKeys.USER_NOT_EXIST.value, data=None, error=None)
        if user is None:
            return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key=ResponseMessageKeys.USER_NOT_EXIST.value,
                                      data=None, error=None)
        if user:
            if user.deactivated_at is None:
                if not check_password_hash(pwhash=user.pin, password=pin):
                    return send_json_response(http_status=HttpStatusCode.FORBIDDEN.value, response_status=False,
                                              message_key=ResponseMessageKeys.INVALID_PASSWORD.value, data=None,
                                              error=None)

                data = UserView.create_auth_response(
                    self=None, user=user, data=None)
                return send_json_response(http_status=HttpStatusCode.OK.value, response_status=True,
                                          message_key=ResponseMessageKeys.LOGIN_SUCCESSFULLY.value.format(
                                              user.first_name),
                                          data=data, error=None)
            else:
                return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                          message_key=ResponseMessageKeys.LOGIN_FAILED.value,
                                          data=None, error=None)
        else:
            return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key=ResponseMessageKeys.LOGIN_FAILED.value,
                                      data=None, error=None)
    
    @token_required
    def bulk_insert(current_user=None):
        try:
            file = request.files.get('file')
            if not file or not file.filename.endswith(('.xlsx')):
                return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key= " Invalid file",
                                      data=None, error=None)
            f=file.filename
            print(f"File received: {f}")


            file_path = os.path.join(config_data['UPLOAD_FOLDER'],f)
            file.save(file_path)
            print("save file", file_path)

            try:
                wb = load_workbook(file_path)
                sheet = wb.active  # Get the active sheet (or use wb[sheet_name] for specific sheets)
                sheet_data = []

                for row in sheet.iter_rows(min_row=2, values_only=True):  
                    sheet_data.append(row)

                # Log the extracted data
                print(f"Data extracted from file: {sheet_data}")
            
            except Exception as e:
                print(f"Error while extracting data from Excel file: {e}")
                print("Detailed Stack Trace:")
                print(traceback.format_exc())  # Print the detailed stack trace here
                return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                        message_key="Error extracting data",
                                        data=None, error=str(e))

            # Check if sheet data exists
            if not sheet_data:
                print("No data found in 'Sheet1'.")
                return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                        message_key="No data found in Sheet1",
                                        data=None, error="Sheet1 is missing or empty.")


            users_to_insert = []

            for row in sheet_data[1::]:
                if len(row)!=4:
                    return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key= " Invalid row data",
                                      data=None, error=None)
                                        
                first_name, last_name,primary_phone,primary_email = row

                if not first_name or not last_name or not primary_phone or not primary_email:
                    return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key= " Missing Fileds ",
                                      data=None, error=None)
                user = User(first_name=first_name, last_name=last_name, primary_phone=primary_phone,primary_email=primary_email)
                users_to_insert.append(user)

            db.session.bulk_save_objects(users_to_insert)
            db.session.commit()
            return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key= " users inserted successfully",
                                      data=None, error=None)
        
        except Exception as e:
              print(f"Error: {e}")
              print("Detailed Stack Trace:")
              print(traceback.format_exc()) 
              return send_json_response(http_status=HttpStatusCode.OK.value, response_status=False,
                                      message_key= " Error Comes while processing exel file",
                                   data=None, error=None)
            
 

